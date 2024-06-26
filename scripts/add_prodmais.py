from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from dotenv import load_dotenv
import os
import openpyxl

load_dotenv()

# browser = webdriver.Firefox()
browser = webdriver.Firefox(executable_path='C:\\geckodriver.exe')

url = 'http://localhost:8080/inclusao.php'
username = os.getenv('PRODMAIS_USERNAME')
password = os.getenv('PRODMAIS_PASSWORD')

browser.get(url)

start_time = time.time()

browser.find_element(By.NAME, 'username').send_keys(username)
browser.find_element(By.NAME, 'password').send_keys(password)
time.sleep(0.5)
browser.find_element(By.NAME, 'submit').click()
time.sleep(1)

# cvs_dir = '/home/pedrodl/Documents/ProdMaisInsper/ProdMais-Utils/curriculos'
cvs_dir = 'C:\\Users\\Yan\\Desktop\\utilsProdmais\\ProdMais-Utils\\curriculos'
c = 1
# local dos docentes é o mesmo do arquivo atual + dados_docentes.csv
dados_docentes_dir = 'data\\dados_docentes.xlsx'
stats_dir = 'data\\status_add.xlsx'

try:
    # Carregue o arquivo de workbook apenas uma vez
    workbook = openpyxl.load_workbook(dados_docentes_dir)
    sheet = workbook['docentes_lattes']
    
    workbook_stats = openpyxl.load_workbook(stats_dir)
    stats_sheet = workbook_stats[workbook_stats.sheetnames[0]]

    for row in range(2, sheet.max_row + 1):
        url_lattes = sheet.cell(row=row, column=3).value
        id = url_lattes.split('/')[-1]

        for filename in os.listdir(cvs_dir):
            if filename.endswith('.xml') and id in filename:
                file_path = os.path.join(cvs_dir, filename)
                
                browser.find_element(By.ID, 'fileXML').send_keys(file_path)
                time.sleep(1)

                # encontra campo name="genero" e digita o conteúdo da coluna 12
                genero = sheet.cell(row=row, column=12).value
                browser.find_element(By.NAME, 'genero').clear()
                
                # se genero != None, digite genero
                if genero:
                    browser.find_element(By.NAME, 'genero').send_keys(genero)
                time.sleep(0.05)
                
                google_scholar = sheet.cell(row=row, column=13).value
                browser.find_element(By.NAME, 'google_citation').clear()
                if google_scholar:
                    browser.find_element(By.NAME, 'google_citation').send_keys(google_scholar)
                    
                browser.find_element(By.NAME, 'lattes_id').clear()
                browser.find_element(By.NAME, 'lattes_id').send_keys(id)
                time.sleep(0.05)
                
                linkedin = sheet.cell(row=row, column=22).value
                browser.find_element(By.NAME, 'linkedin').clear()
                if linkedin:
                    browser.find_element(By.NAME, 'linkedin').send_keys(linkedin)
                time.sleep(0.05)
                    
                unidades_academicas = []
                direito = sheet.cell(row=row, column=19).value
                negocios = sheet.cell(row=row, column=20).value
                tecnologia = sheet.cell(row=row, column=21).value
                
                if direito:
                    unidades_academicas.append('Direito')
                if negocios:
                    unidades_academicas.append('Negócios')
                if tecnologia:
                    unidades_academicas.append('Tecnologia')
                unidade_academica = '|'.join(unidades_academicas)
                
                
                browser.find_element(By.NAME, 'unidade_academica').clear()
                if unidade_academica:
                    browser.find_element(By.NAME, 'unidade_academica').send_keys(unidade_academica)
                    time.sleep(0.05)
                
                dot = sheet.cell(row=row, column=14).value
                mpa = sheet.cell(row=row, column=15).value
                mpe = sheet.cell(row=row, column=16).value
                mpp = sheet.cell(row=row, column=17).value
                dpa = sheet.cell(row=row, column=18).value
                
                ppgs = []
                if dot:
                    ppgs.append('Doutorado Acadêmico em Economia')  #precisa confirmar se esse é o nome do DOT
                if mpa:
                    ppgs.append('Mestrado Profissional em Administração')
                if mpe:
                    ppgs.append('Mestrado Profissional em Economia')
                if mpp:
                    ppgs.append('Mestrado Profissional em Políticas Públicas')
                if dpa:
                    ppgs.append('Doutorado Profissional em Administração')
                
                ppgs = '|'.join(ppgs)
                browser.find_element(By.NAME, 'ppg_nome').clear()
                if ppgs:
                    browser.find_element(By.NAME, 'ppg_nome').send_keys(ppgs)
                
                tipo_vinculo = str(sheet.cell(row=row, column=11).value)
                browser.find_element(By.NAME, 'tipvin').clear()
                if tipo_vinculo in ['0']:
                    browser.find_element(By.NAME, 'tipvin').send_keys("Tempo Parcial")
                if tipo_vinculo in ['1', '2', '3']:
                    browser.find_element(By.NAME, 'tipvin').send_keys("Tempo Integral")
                time.sleep(1)
                
                browser.find_element(By.XPATH, '/html/body/main/div/div/form[1]/div[2]/button').click()
                time.sleep(2)

                page_content = browser.page_source
                if ('Registro anterior não encontrado na base') in page_content:
                    print(f"Arquivo {filename} [{c}] adicionado")
                    stats_sheet.cell(row=row, column=9).value = 'OK'
                    stats_sheet.cell(row=row, column=10).value = time.strftime('%H:%M:%S %d/%m/%Y')
                else:
                    print(f"Arquivo {filename} [{c}] já existe na base")
                    stats_sheet.cell(row=row, column=8).value = 'Já existe'
                    stats_sheet.cell(row=row, column=9).value = time.strftime('%H:%M:%S %d/%m/%Y')
                c+=1

                browser.back()
                time.sleep(2)

                # Salve as alterações a cada iteração
                workbook_stats.save(stats_dir)
except Exception as e:
    print(f"An error occurred: {str(e)}")
finally:
    end_time = time.time()
    final_time = (end_time - start_time)/60

    print(f"Tempo de execução do script: {final_time:2f} minutos")

    browser.quit()
