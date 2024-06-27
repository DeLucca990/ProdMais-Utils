from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from dotenv import load_dotenv
import os
import openpyxl
from datetime import datetime

load_dotenv()

browser = webdriver.Firefox()
# browser = webdriver.Firefox(executable_path='C:\\geckodriver.exe')

url = 'http://localhost:8080/inclusao.php'
username = os.getenv('PRODMAIS_USERNAME')
password = os.getenv('PRODMAIS_PASSWORD')

browser.get(url)

start_time = time.time()

browser.find_element(By.NAME, 'username').send_keys(username)
browser.find_element(By.NAME, 'password').send_keys(password)
time.sleep(0.5)
browser.find_element(By.NAME, 'submit').click()
time.sleep(1)


# dados_ppg = 'data\\DadosGeraisStricto.xlsx'
dados_ppg = '/home/pedrodl/Documents/ProdMaisInsper/ProdMais-Utils/data/DadosGeraisStricto.xlsx'

try:
    # Carregue o arquivo de workbook apenas uma vez
    workbook = openpyxl.load_workbook(dados_ppg)
    sheet = workbook['Info.geral']

    for row in range(2, sheet.max_row + 1):
        id_curso = sheet.cell(row=row, column=1).value
        codigo_capes = sheet.cell(row=row, column=2).value
        conceito_capes = sheet.cell(row=row, column=3).value
        instituicao = sheet.cell(row=row, column=4).value
        nome_ppg = sheet.cell(row=row, column=8).value
        data_inicio = sheet.cell(row=row, column=9).value
        site_ppg = sheet.cell(row=row, column=10).value
        email_ppg = sheet.cell(row=row, column=11).value
        nome_coordenador = sheet.cell(row=row, column=13).value
        data_inicio_coordenador = sheet.cell(row=row, column=14).value
        nivel = sheet.cell(row=row, column=15).value
        url_dataverse = sheet.cell(row=row, column=16).value
        
        print(f"ID_CURSO: {id_curso}")
        print(f"COD_CAPES: {codigo_capes}")
        print(f"CONCEITO_CAPES: {conceito_capes}")
        print(f"NOME_INSTITUICAO: {instituicao}")
        print(f"NOME_PPG: {nome_ppg}")
        print(f"INI_PPG: {data_inicio}")
        print(f"PPG_SITE: {site_ppg}")
        print(f"PPG_EMAIL: {email_ppg}")
        print(f"NOME_COORDENADOR: {nome_coordenador}")
        print(f"DT_INI_COORD: {data_inicio_coordenador}")
        print(f"NIVEL: {nivel}")
        print(f"URL_DATAVERSE: {url_dataverse}")
        

        browser.find_element(By.NAME, 'ID_CURSO').clear()
        if id_curso:
            browser.find_element(By.NAME, 'ID_CURSO').send_keys(id_curso)
        print(f"id curso:{id_curso}")
        browser.find_element(By.NAME, 'COD_CAPES').clear()
        if codigo_capes:
            browser.find_element(By.NAME, 'COD_CAPES').send_keys(codigo_capes)
        
        browser.find_element(By.NAME, 'CONCEITO_CAPES').clear()
        if conceito_capes:
            browser.find_element(By.NAME, 'CONCEITO_CAPES').send_keys(conceito_capes)

        browser.find_element(By.NAME, 'NOME_INSTITUICAO').clear()
        if instituicao:
            browser.find_element(By.NAME, 'NOME_INSTITUICAO').send_keys(instituicao)
        
        browser.find_element(By.NAME, 'NOME_PPG').clear()
        if nome_ppg:
            browser.find_element(By.NAME, 'NOME_PPG').send_keys(nome_ppg)
            
            
        browser.find_element(By.NAME, 'INI_PPG').clear()
        if data_inicio:
            browser.find_element(By.NAME, 'INI_PPG').clear()
            if data_inicio:
                if(type(data_inicio) == str):
                    browser.find_element(By.NAME, 'INI_PPG').send_keys(data_inicio)
                else:
                    formatted_date = datetime.strftime(data_inicio, "%d/%m/%Y")
                    browser.find_element(By.NAME, 'INI_PPG').send_keys(formatted_date)
        
        browser.find_element(By.NAME, 'PPG_SITE').clear()
        if site_ppg:
            browser.find_element(By.NAME, 'PPG_SITE').send_keys(site_ppg)
        
        browser.find_element(By.NAME, 'PPG_EMAIL').clear()
        if email_ppg:
            browser.find_element(By.NAME, 'PPG_EMAIL').send_keys(email_ppg)
        
        browser.find_element(By.NAME, 'NOME_COORDENADOR').clear()
        if nome_coordenador:
            browser.find_element(By.NAME, 'NOME_COORDENADOR').send_keys(nome_coordenador)
        
        browser.find_element(By.NAME, 'DT_INI_COORD').clear()
        if data_inicio_coordenador:
            formatted_date = datetime.strftime(data_inicio_coordenador, "%d/%m/%Y")
            browser.find_element(By.NAME, 'DT_INI_COORD').send_keys(formatted_date)
            
        browser.find_element(By.NAME, 'NIVEL').clear()
        if nivel:
            browser.find_element(By.NAME, 'NIVEL').send_keys(nivel)
        
        browser.find_element(By.NAME, 'NOME_CAMPUS').clear()
        browser.find_element(By.NAME, 'NOME_CAMPUS').send_keys('São Paulo')
        
        browser.find_element(By.NAME, 'PRODMAIS_DATAVERSE').clear()
        if url_dataverse:
            browser.find_element(By.NAME, 'PRODMAIS_DATAVERSE').send_keys(url_dataverse)

        browser.find_element(By.XPATH, '/html/body/main/div/div/form[2]/div[2]/button').click()
        browser.back()
        time.sleep(2)
                
except Exception as e:
    print(f"An error occurred: {str(e)}")
finally:
    end_time = time.time()
    final_time = (end_time - start_time)/60

    print(f"Tempo de execução do script: {final_time:2f} minutos")

    browser.quit()
