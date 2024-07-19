from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import time
from dotenv import load_dotenv
import os
import openpyxl
from selenium.webdriver.support.ui import WebDriverWait

load_dotenv()

# # Parâmetros Pedro
# url = 'http://localhost:8080/inclusao.php'
# browser = webdriver.Firefox()
# cvs_dir = '/home/pedrodl/Documents/ProdMaisInsper/ProdMais-Utils/curriculos'
# dados_docentes_dir = '/home/pedrodl/Documents/ProdMaisInsper/ProdMais-Utils/data/dados_docentes.xlsx'
# stats_dir = '/home/pedrodl/Documents/ProdMaisInsper/ProdMais-Utils/data/status_add.xlsx'

# Parâmetros Yan
options = webdriver.FirefoxOptions()
options.headless = False
url = 'https://prodmais.datascience.insper.edu.br/inclusao.php'
browser = webdriver.Firefox(executable_path='C:\\geckodriver.exe', options=options)
cvs_dir = 'C:\\Users\\Yan\\Desktop\\utilsProdmais\\ProdMais-Utils\\curriculos'
dados_docentes_dir = 'data\\dados_docentes.xlsx'
stats_dir = 'data\\status_add.xlsx' 

username = os.getenv('PRODMAIS_USERNAME')
password = os.getenv('PRODMAIS_PASSWORD')

browser.get(url)

start_time = time.time()

browser.find_element(By.NAME, 'username').send_keys(username)
browser.find_element(By.NAME, 'password').send_keys(password)
time.sleep(0.5)
browser.find_element(By.NAME, 'submit').click()
time.sleep(1)

c = 1
try:
    # Carregue o arquivo de workbook apenas uma vez
    workbook = openpyxl.load_workbook(dados_docentes_dir)
    sheet = workbook['docentes_lattes']
    
    workbook_stats = openpyxl.load_workbook(stats_dir)
    stats_sheet = workbook_stats[workbook_stats.sheetnames[0]]

    # rows_falha = [43, 133, 193, 285, 324, 381, 391, 406, 469]
    for row in range(2, sheet.max_row + 1):
    # for row in rows_falha:
        url_lattes = sheet.cell(row=row, column=3).value
        id = url_lattes.split('/')[-1]

        for filename in os.listdir(cvs_dir):
            if filename.endswith('.xml') and id in filename:
                file_path = os.path.join(cvs_dir, filename)
                
                browser.find_element(By.ID, 'fileXML').send_keys(file_path)
                time.sleep(1)

                # encontra campo name="genero" e digita o conteúdo da coluna 12
                genero = sheet.cell(row=row, column=12).value
                browser.find_element(By.NAME, 'genero').clear()
                
                # se genero != None, digite genero
                if genero:
                    browser.find_element(By.NAME, 'genero').send_keys(genero)
                time.sleep(0.05)
                
                google_scholar = sheet.cell(row=row, column=13).value
                browser.find_element(By.NAME, 'google_citation').clear()
                time.sleep(0.005)
                if google_scholar:
                    browser.find_element(By.NAME, 'google_citation').send_keys(google_scholar)
                    
                browser.find_element(By.NAME, 'lattes_id').clear()
                time.sleep(0.005)
                browser.find_element(By.NAME, 'lattes_id').send_keys(id)
                time.sleep(0.05)
                
                linkedin = sheet.cell(row=row, column=22).value
                browser.find_element(By.NAME, 'linkedin').clear()
                time.sleep(0.005)
                if linkedin:
                    browser.find_element(By.NAME, 'linkedin').send_keys(linkedin)
                time.sleep(0.05)
                    
                unidades_academicas = []
                direito = sheet.cell(row=row, column=19).value
                negocios = sheet.cell(row=row, column=20).value
                tecnologia = sheet.cell(row=row, column=21).value
                
                if direito:
                    unidades_academicas.append('Direito')
                if negocios:
                    unidades_academicas.append('Negócios')
                if tecnologia:
                    unidades_academicas.append('Tecnologia')
                unidade_academica =  ''
                unidade_academica = '|'.join(unidades_academicas)
                
                
                browser.find_element(By.NAME, 'unidade_academica').clear()
                time.sleep(0.005)
                if unidade_academica:
                    browser.find_element(By.NAME, 'unidade_academica').send_keys(unidade_academica)
                    time.sleep(0.05)
                else:
                    browser.find_element(By.NAME, 'unidade_academica').send_keys('Não Classificado')
                    time.sleep(0.05)
                
                dot = sheet.cell(row=row, column=14).value
                mpa = sheet.cell(row=row, column=15).value
                mpe = sheet.cell(row=row, column=16).value
                mpp = sheet.cell(row=row, column=17).value
                dpa = sheet.cell(row=row, column=18).value
                
                ppgs = []
                if dot:
                    ppgs.append('Doutorado Acadêmico em Economia')  #precisa confirmar se esse é o nome do DOT
                if mpa:
                    ppgs.append('Mestrado Profissional em Administração')
                if mpe:
                    ppgs.append('Mestrado Profissional em Economia')
                if mpp:
                    ppgs.append('Mestrado Profissional em Políticas Públicas')
                if dpa:
                    ppgs.append('Doutorado Profissional em Administração')
                
                ppgs = '|'.join(ppgs)
                ppg_nome_field = browser.find_element(By.NAME, 'ppg_nome')
                ppg_nome_field.clear()
                assert ppg_nome_field.get_attribute("value") == ""
                time.sleep(0.05)
                ppg_nome_field.clear()
                if ppgs:
                    browser.find_element(By.NAME, 'ppg_nome').send_keys(ppgs)
                else:
                    browser.find_element(By.NAME, 'ppg_nome').send_keys('Não Classificado')
                
                tipo_vinculo = str(sheet.cell(row=row, column=11).value)
                browser.find_element(By.NAME, 'tipvin').clear()
                time.sleep(0.005)
                if tipo_vinculo == "TI":
                    browser.find_element(By.NAME, 'tipvin').send_keys("Tempo Integral")
                elif tipo_vinculo == 'TP':
                    browser.find_element(By.NAME, 'tipvin').send_keys("Tempo Parcial")
                else:
                    browser.find_element(By.NAME, 'tipvin').send_keys("Outros")
                time.sleep(0.05)
                
                email_institucional = sheet.cell(row=row, column=23).value
                browser.find_element(By.NAME, 'email').clear()
                if email_institucional:
                    browser.find_element(By.NAME, 'email').send_keys(email_institucional)
                time.sleep(0.05)
                
                raca = sheet.cell(row=row, column=24).value
                browser.find_element(By.NAME, 'raca').clear()
                if raca:
                    browser.find_element(By.NAME, 'raca').send_keys(raca)
                time.sleep(0.05)
                
                site_pessoal = sheet.cell(row=row, column=28).value
                browser.find_element(By.NAME, 'site_pessoal').clear()
                if site_pessoal:
                    browser.find_element(By.NAME, 'site_pessoal').send_keys(site_pessoal)
                time.sleep(0.05)
                    
                areas = sheet.cell(row=row, column=25).value
                areas = areas.split('|')
                areas_formatadas = []
                for area in areas:
                    if "-" in area:
                        area = area.split("-")[1]
                        if area.strip() != "" and 'Outra(s) área(s)' not in area:
                            areas_formatadas.append(area.strip())
                    else:
                        if area.strip() != "" and area.strip() != 'Outra(s) área(s)':
                            areas_formatadas.append(area.strip())
                areas_formatadas = '|'.join(areas_formatadas)
                browser.find_element(By.NAME, 'areas').clear()
                if len(areas_formatadas) > 0:
                    browser.find_element(By.NAME, 'areas').send_keys(areas_formatadas)
                
                programas = sheet.cell(row=row, column=26).value
                browser.find_element(By.NAME, 'programas').clear()
                if len(programas) > 0:
                    browser.find_element(By.NAME, 'programas').send_keys(programas)
                else:
                    browser.find_element(By.NAME, 'programas').send_keys('NULL')
                
                industrias = sheet.cell(row=row, column=27).value
                if industrias:
                    industrias = industrias.replace("Setor de ", "").replace("Setor ", "")
                    industrias = industrias.replace(" |", "|").replace("| ", "|")
                browser.find_element(By.NAME, 'industrias').clear()
                if len(industrias) > 0:
                    browser.find_element(By.NAME, 'industrias').send_keys(industrias)    
                
                imprensa_temas = sheet.cell(row=row, column=29).value
                lista_temas = []
                if imprensa_temas:
                    lista_temas = imprensa_temas.split(',')
                    #retira espaços em branco da lista
                    lista_temas = [tema.strip() for tema in lista_temas]
                    #tira elementos vazios da lista
                    lista_temas = list(filter(None, lista_temas))
                temas = '|'.join(lista_temas)
                browser.find_element(By.NAME, 'imprensa_tema').clear()
                if len(temas) > 0:
                    browser.find_element(By.NAME, 'imprensa_tema').send_keys(temas)
                else:
                    browser.find_element(By.NAME, 'imprensa_tema').send_keys('Não Classificado')
                
                browser.find_element(By.XPATH, '/html/body/main/div/div/form[1]/div[2]/button').click()
                
                wait = WebDriverWait(browser, 120)  # 120 segundos de timeout

                # Função personalizada para verificar ambos textos
                def text_to_be_present_in_element(locator, text1, text2, text3, text4, text5):
                    def _predicate(driver):
                        try:
                            element = driver.find_element(*locator)
                            if text1 in element.text or text2 in element.text or text3 in element.text or text4 in element.text or text5 in element.text:
                                return True
                        except:
                            return False
                    return _predicate

                # Localizador do elemento que contém o texto (ajuste conforme necessário)
                locator = (By.TAG_NAME, 'body')  # Pode ser By.CLASS_NAME, By.ID, etc.

                # Esperar até que o texto esteja presente no elemento
                wait.until(text_to_be_present_in_element(locator, "Tem", "Registro anterior não encontrado na base", "413 Request Entity Too Large", "Nenhum arquivo foi enviado", "Ocorreu um erro ao enviar o arquivo"))
                
                #caso exista 413 Request Entity Too Large, printa o erro e pula para o próximo arquivo
                
                if('413 Request Entity Too Large' in browser.page_source):
                    print(f"Arquivo {filename} [{c}] muito grande")
                    stats_sheet.cell(row=row, column=8).value = '413 Request Entity Too Large'
                    stats_sheet.cell(row=row, column=9).value = time.strftime('%H:%M:%S %d/%m/%Y')
                    c+=1
                    browser.back()
                    WebDriverWait(browser, 120).until(EC.presence_of_element_located((By.NAME, 'NOME_INSTITUICAO')))
                    workbook_stats.save(stats_dir)
                    continue
                
                if('Nenhum arquivo foi enviado' in browser.page_source):
                    print(f"Arquivo {filename} [{c}] não enviado")
                    stats_sheet.cell(row=row, column=8).value = 'Nenhum arquivo foi enviado'
                    stats_sheet.cell(row=row, column=9).value = time.strftime('%H:%M:%S %d/%m/%Y')
                    c+=1
                    browser.back()
                    WebDriverWait(browser, 120).until(EC.presence_of_element_located((By.NAME, 'NOME_INSTITUICAO')))
                    workbook_stats.save(stats_dir)
                    continue
                    
                if('Ocorreu um erro ao enviar o arquivo' in browser.page_source):
                    print(f"Arquivo {filename} [{c}] não enviado")
                    stats_sheet.cell(row=row, column=8).value = 'Ocorreu um erro ao enviar o arquivo'
                    stats_sheet.cell(row=row, column=9).value = time.strftime('%H:%M:%S %d/%m/%Y')
                    c+=1
                    browser.back()
                    WebDriverWait(browser, 120).until(EC.presence_of_element_located((By.NAME, 'NOME_INSTITUICAO')))
                    workbook_stats.save(stats_dir)
                    continue
                
                page_content = browser.page_source
                if ('Registro anterior não encontrado na base') in page_content:
                    print(f"Arquivo {filename} [{c}] adicionado")
                    stats_sheet.cell(row=row, column=8).value = 'OK'
                    stats_sheet.cell(row=row, column=9).value = time.strftime('%H:%M:%S %d/%m/%Y')
                else:
                    print(f"Arquivo {filename} [{c}] já existe na base")
                    stats_sheet.cell(row=row, column=8).value = 'Já existe'
                    stats_sheet.cell(row=row, column=9).value = time.strftime('%H:%M:%S %d/%m/%Y')
                c+=1

                browser.back()
                #espera haver o campo com name instituição
                WebDriverWait(browser, 120).until(EC.presence_of_element_located((By.NAME, 'NOME_INSTITUICAO')))

                # Salve as alterações a cada iteração
                workbook_stats.save(stats_dir)
except Exception as e:
    print(f"An error occurred: {str(e)}")
finally:
    end_time = time.time()
    final_time = (end_time - start_time)/60

    print(f"Tempo de execução do script: {final_time:.2f} minutos")

    browser.quit()
