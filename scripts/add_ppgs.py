from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from dotenv import load_dotenv
import os
import openpyxl

load_dotenv()

# browser = webdriver.Firefox()
browser = webdriver.Firefox(executable_path='C:\\geckodriver.exe')

url = 'http://localhost:8080/inclusao.php'
username = os.getenv('PRODMAIS_USERNAME')
password = os.getenv('PRODMAIS_PASSWORD')

browser.get(url)

start_time = time.time()

browser.find_element(By.NAME, 'username').send_keys(username)
browser.find_element(By.NAME, 'password').send_keys(password)
time.sleep(0.5)
browser.find_element(By.NAME, 'submit').click()
time.sleep(1)


dados_ppg = 'data\\DadosGeraisStricto.xlsx'

try:
    # Carregue o arquivo de workbook apenas uma vez
    workbook = openpyxl.load_workbook(dados_ppg)
    sheet = workbook['Planilha1']

    for row in range(2, sheet.max_row + 1):
        id_curso = sheet.cell(row=row, column=1).value
        codigo_capes = sheet.cell(row=row, column=2).value
        conceito_capes = sheet.cell(row=row, column=3).value
        instituicao = sheet.cell(row=row, column=4).value
        nome_ppg = sheet.cell(row=row, column=7).value
        data_inicio = sheet.cell(row=row, column=8).value
        site_ppg = sheet.cell(row=row, column=9).value
        email_ppg = sheet.cell(row=row, column=10).value
        nome_coordenador = sheet.cell(row=row, column=12).value
        data_inicio_coordenador = sheet.cell(row=row, column=13).value
        nivel = sheet.cell(row=row, column=14).value

        browser.find_element(By.NAME, 'ID_CURSO').clear()
        if id_curso:
            browser.find_element(By.NAME, 'ID_CURSO').send_keys(id_curso)
        
        browser.find_element(By.NAME, 'COD_CAPES').clear()
        if codigo_capes:
            browser.find_element(By.NAME, 'COD_CAPES').send_keys(codigo_capes)
        
        browser.find_element(By.NAME, 'CONCEITO_CAPES').clear()
        if conceito_capes:
            browser.find_element(By.NAME, 'CONCEITO_CAPES').send_keys(conceito_capes)

        browser.find_element(By.NAME, 'NOME_INSTITUICAO').clear()
        if instituicao:
            browser.find_element(By.NAME, 'NOME_INSTITUICAO').send_keys(instituicao)
        
        browser.find_element(By.NAME, 'NOME_PPG').clear()
        if nome_ppg:
            browser.find_element(By.NAME, 'NOME_PPG').send_keys(nome_ppg)
            
            
        browser.find_element(By.NAME, 'INI_PPG').clear()
        if data_inicio:
            browser.find_element(By.NAME, 'INI_PPG').send_keys(data_inicio)
        
        browser.find_element(By.NAME, 'PPG_SITE').clear()
        if site_ppg:
            browser.find_element(By.NAME, 'PPG_SITE').send_keys(site_ppg)
        
        browser.find_element(By.NAME, 'PPG_EMAIL').clear()
        if email_ppg:
            browser.find_element(By.NAME, 'PPG_EMAIL').send_keys(email_ppg)
        
        browser.find_element(By.NAME, 'NOME_COORDENADOR').clear()
        if nome_coordenador:
            browser.find_element(By.NAME, 'NOME_COORDENADOR').send_keys(nome_coordenador)
        
        browser.find_element(By.NAME, 'DT_INI_COORD').clear()
        if data_inicio_coordenador:
            browser.find_element(By.NAME, 'DT_INI_COORD').send_keys(data_inicio_coordenador)
            
        browser.find_element(By.NAME, 'NIVEL').clear()
        if nivel:
            browser.find_element(By.NAME, 'NIVEL').send_keys(nivel)

        browser.back()
        time.sleep(1)
                
except Exception as e:
    print(f"An error occurred: {str(e)}")
finally:
    end_time = time.time()
    final_time = (end_time - start_time)/60

    print(f"Tempo de execução do script: {final_time:2f} minutos")

    browser.quit()
